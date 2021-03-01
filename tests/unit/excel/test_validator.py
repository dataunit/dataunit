import unittest
import tempfile
import openpyxl
from openpyxl import Workbook
from dataunit.excel.validator import validate_workbook
from dataunit.excel.constants import REQUIRED_COLUMNS, REQUIRED_SHEETS

def generate_test_workbook():
    # helper method to generate test workbook with required sheets and columns

    workbook = openpyxl.Workbook()

    for sheet_name in REQUIRED_SHEETS:
        sheet = workbook.create_sheet(sheet_name)

        col = 1
        for column_name in REQUIRED_COLUMNS[sheet_name]:
            sheet.cell(1, col).value = column_name
            col = col + 1

    return workbook
    

class TestValidator(unittest.TestCase):

    def test_with_valid_workbook(self):
        workbook = generate_test_workbook()
        with tempfile.TemporaryDirectory() as tmpdirname:
            test_workbook = tmpdirname + "/empty_workbook.xlsx"            
            workbook.save(test_workbook)  

            #Note: there is no assertion here as this should be valid
            try:
                validate_workbook(test_workbook)
            except:
                self.fail('Method call should not raise exception.')


    def test_validation_with_empty_workbook(self):

        #create empty workbook
        workbook = openpyxl.Workbook()
        with tempfile.TemporaryDirectory() as tmpdirname:
            test_workbook = tmpdirname + "/empty_workbook.xlsx"            
            workbook.save(test_workbook)  

            with self.assertRaises(AssertionError) as context:
                validate_workbook(test_workbook)

            self.assertTrue('Required sheet "Tests" not found in workbook ' in str(context.exception))

    def test_validation_with_missing_settings_sheet(self):

        workbook = openpyxl.Workbook()

        tests_sheet = workbook.create_sheet('Tests')

        # get list of columns for the Tests sheet and add them
        columns = REQUIRED_COLUMNS['Tests']
        col = 1
        for column_name in columns:
            tests_sheet.cell(1, col).value = column_name
            col = col + 1

        with tempfile.TemporaryDirectory() as tmpdirname:
            test_workbook = tmpdirname + "/empty_workbook.xlsx"            
            workbook.save(test_workbook)

            with self.assertRaises(AssertionError) as context:        
                validate_workbook(test_workbook)

            self.assertTrue('Required sheet "Settings" not found in workbook ' in str(context.exception))

    def test_validation_with_missing_test_name_column(self):

        workbook = generate_test_workbook()
        workbook['Tests']["A1"] = "Missing"
        with tempfile.TemporaryDirectory() as tmpdirname:
            test_workbook = tmpdirname + "/empty_workbook.xlsx"            
            workbook.save(test_workbook)

            with self.assertRaises(AssertionError) as context:        
                validate_workbook(test_workbook)
            self.assertTrue('Required column name "Test Name" not found in workbook ' in str(context.exception))

            